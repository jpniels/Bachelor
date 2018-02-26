import time
import openpyxl
import xlrd
import MySQLdb
print ("hej")
from openpyxl import load_workbook
book = xlrd.open_workbook("C:\Users\Jonas\Desktop\Data_pjat2.xlsx")
print (book.sheet_names())

sheet = book.sheet_by_name("Ark2")
print ("opening file")

print ("after sheet")

database = MySQLdb.connect( host = "localhost", user = "root", passwd = "", db = "bachelor")
print ("connected")
cursor = database.cursor()

query  = """INSERT INTO data (Floor, Room, SensorType, Modality, Unit, Reading, TimeInMiliseconds) VAlUES (%s,%s,%s,%s,%s,%s,%s)"""


for row in range(1, sheet.nrows):
    Floor             = sheet.cell(row,0).value
    Room              = sheet.cell(row,1).value
    SensorType        = sheet.cell(row,2).value
    Modality          = sheet.cell(row,3).value
    Unit              = sheet.cell(row,4).value
    Reading           = sheet.cell(row,6).value
    TimeInMiliseconds = sheet.cell(row,8).value

    values = (Floor, Room, SensorType, Modality, Unit, Reading, TimeInMiliseconds)

    cursor.execute(query, values)

cursor.close()

database.commit()

database.close()